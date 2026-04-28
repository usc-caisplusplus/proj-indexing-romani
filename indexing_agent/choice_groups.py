"""Load the indexing-term choice groups from the Excel workbook and
build / cache OpenAI embeddings for every label so we can run nearest-
neighbour lookups (the RAG step that maps free-form transcript text to
controlled-vocabulary keywords).
"""

from __future__ import annotations

import hashlib
import json
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import openpyxl
from openai import OpenAI
from tenacity import retry, stop_after_attempt, wait_random_exponential
from tqdm import tqdm

logger = logging.getLogger(__name__)


# Mapping from the canonical "field key" used everywhere in this codebase
# to the QuestionID sheet that holds its choice group.
FIELD_TO_QUESTION_SHEET: Dict[str, str] = {
    "city_of_birth": "QuestionID 1806 choices",
    "country_of_birth": "QuestionID 1807 choices",
    "religious_identity_prewar": "QuestionID 1816 choices",
    "religious_identity_postwar": "QuestionID 1817 choices",
    "ghettos": "QuestionID 1809 choices",
    "camps": "QuestionID 1808 choices",
    "hiding_location": "QuestionID 3603 choices",
    "hiding_type": "QuestionID 3604 choices",
    "resistance_groups": "QuestionID 3605 choices",
    "who_liberated": "QuestionID 3606 choices",
    "where_liberated": "QuestionID 3607 choices",
    "other_attributes": "QuestionID 1813 choices",
    "special_events": "QuestionID 1814 choices",
}

# Whether new terms can be proposed for a given field (per guideline 2.5.17 -
# Special Events disallows new term proposals).
FIELD_ALLOWS_PROPOSED: Dict[str, bool] = {
    "city_of_birth": True,
    "country_of_birth": True,
    "religious_identity_prewar": True,
    "religious_identity_postwar": True,
    "ghettos": True,
    "camps": True,
    "hiding_location": True,
    "hiding_type": True,
    "resistance_groups": True,
    "who_liberated": True,
    "where_liberated": True,
    "other_attributes": True,
    "special_events": False,
}


@dataclass
class ChoiceGroup:
    field: str
    sheet: str
    keyword_ids: np.ndarray  # shape (N,)
    labels: List[str]        # length N
    embeddings: np.ndarray   # shape (N, D), L2-normalised
    embed_model: str

    def search(self, query_vec: np.ndarray, top_k: int = 15) -> List[Tuple[int, str, float]]:
        """Cosine-similarity search.  `query_vec` must be L2-normalised."""
        scores = self.embeddings @ query_vec
        top_idx = np.argpartition(-scores, min(top_k, len(scores) - 1))[:top_k]
        top_idx = top_idx[np.argsort(-scores[top_idx])]
        return [(int(self.keyword_ids[i]), self.labels[i], float(scores[i])) for i in top_idx]


# ---------------------------------------------------------------------------
# Loading from Excel
# ---------------------------------------------------------------------------


def _load_sheet_labels(xlsx_path: Path, sheet_name: str) -> Tuple[List[int], List[str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet {sheet_name!r} not found in {xlsx_path}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    if header is None:
        return [], []
    try:
        kid_col = header.index("KeywordID")
        label_col = header.index("Label")
    except ValueError as exc:  # pragma: no cover - defensive
        raise RuntimeError(f"Unexpected header in {sheet_name}: {header}") from exc

    seen: set[Tuple[int, str]] = set()
    keyword_ids: List[int] = []
    labels: List[str] = []
    for row in rows:
        if row is None:
            continue
        kid = row[kid_col]
        label = row[label_col]
        if kid is None or label is None:
            continue
        try:
            kid_int = int(kid)
        except (TypeError, ValueError):
            continue
        label_str = str(label).strip()
        if not label_str:
            continue
        key = (kid_int, label_str)
        if key in seen:
            continue
        seen.add(key)
        keyword_ids.append(kid_int)
        labels.append(label_str)
    wb.close()
    return keyword_ids, labels


# ---------------------------------------------------------------------------
# OpenAI embedding helpers (with batching + retry)
# ---------------------------------------------------------------------------


@retry(stop=stop_after_attempt(6), wait=wait_random_exponential(min=1, max=30))
def _embed_batch(client: OpenAI, model: str, texts: Sequence[str]) -> List[List[float]]:
    resp = client.embeddings.create(model=model, input=list(texts))
    return [d.embedding for d in resp.data]


def embed_texts(
    client: OpenAI,
    model: str,
    texts: Sequence[str],
    batch_size: int = 256,
    show_progress: bool = False,
) -> np.ndarray:
    if not texts:
        return np.zeros((0, 1536), dtype=np.float32)
    out: List[List[float]] = []
    iterator = range(0, len(texts), batch_size)
    if show_progress:
        iterator = tqdm(list(iterator), desc=f"embed[{model}]", unit="batch")
    for start in iterator:
        chunk = texts[start : start + batch_size]
        out.extend(_embed_batch(client, model, chunk))
    arr = np.asarray(out, dtype=np.float32)
    norms = np.linalg.norm(arr, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    return arr / norms


# ---------------------------------------------------------------------------
# Cache management
# ---------------------------------------------------------------------------


def _cache_key(xlsx_path: Path, sheet_name: str, model: str) -> str:
    h = hashlib.sha1()
    h.update(str(xlsx_path.resolve()).encode())
    h.update(b"|")
    h.update(sheet_name.encode())
    h.update(b"|")
    h.update(model.encode())
    h.update(b"|")
    h.update(str(xlsx_path.stat().st_mtime_ns).encode())
    return h.hexdigest()[:16]


def _cache_paths(cache_dir: Path, key: str) -> Tuple[Path, Path]:
    return cache_dir / f"{key}.npz", cache_dir / f"{key}.json"


def load_choice_group(
    field: str,
    xlsx_path: Path,
    cache_dir: Path,
    client: OpenAI,
    embed_model: str,
) -> ChoiceGroup:
    sheet = FIELD_TO_QUESTION_SHEET[field]
    cache_dir.mkdir(parents=True, exist_ok=True)
    key = _cache_key(xlsx_path, sheet, embed_model)
    npz_path, meta_path = _cache_paths(cache_dir, key)

    if npz_path.exists() and meta_path.exists():
        meta = json.loads(meta_path.read_text())
        data = np.load(npz_path)
        return ChoiceGroup(
            field=field,
            sheet=sheet,
            keyword_ids=data["keyword_ids"],
            labels=meta["labels"],
            embeddings=data["embeddings"],
            embed_model=embed_model,
        )

    logger.info("Building embedding cache for %s (%s)", field, sheet)
    keyword_ids, labels = _load_sheet_labels(xlsx_path, sheet)
    embeddings = embed_texts(client, embed_model, labels, show_progress=True)

    np.savez_compressed(
        npz_path,
        keyword_ids=np.asarray(keyword_ids, dtype=np.int64),
        embeddings=embeddings,
    )
    meta_path.write_text(
        json.dumps(
            {
                "field": field,
                "sheet": sheet,
                "embed_model": embed_model,
                "labels": labels,
                "xlsx": str(xlsx_path),
            }
        )
    )

    return ChoiceGroup(
        field=field,
        sheet=sheet,
        keyword_ids=np.asarray(keyword_ids, dtype=np.int64),
        labels=labels,
        embeddings=embeddings,
        embed_model=embed_model,
    )


def load_all_choice_groups(
    xlsx_path: Path,
    cache_dir: Path,
    client: OpenAI,
    embed_model: str,
    fields: Optional[Iterable[str]] = None,
) -> Dict[str, ChoiceGroup]:
    if fields is None:
        fields = FIELD_TO_QUESTION_SHEET.keys()
    groups: Dict[str, ChoiceGroup] = {}
    for field in fields:
        groups[field] = load_choice_group(
            field=field,
            xlsx_path=xlsx_path,
            cache_dir=cache_dir,
            client=client,
            embed_model=embed_model,
        )
    return groups
